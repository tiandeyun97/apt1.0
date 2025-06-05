from django.views.generic import CreateView, UpdateView, ListView, DetailView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.shortcuts import render, get_object_or_404, redirect
from django.db import IntegrityError, transaction, connection
from django.core.exceptions import ValidationError
from django.contrib import messages
from django.http import JsonResponse, HttpResponseRedirect, HttpResponse
from .models import Task
from tasks.models import Project, TaskStatus
from organize.models import User
from django.db.models import Q
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime, timedelta
from .permissions import TaskPermission
import logging

# 获取logger
logger = logging.getLogger('task_management')

def decode_unicode_escapes(text):
    """解码字符串中的Unicode转义序列"""
    if not text:
        return text
        
    try:
        # 将Unicode转义序列解码为实际字符
        return bytes(str(text), 'utf-8').decode('unicode_escape')
    except Exception:
        return text

class TaskListView(LoginRequiredMixin, ListView):
    model = Task
    template_name = 'admin/task_management/task/task_list.html'
    context_object_name = 'tasks'
    paginate_by = 20
    
    def get_queryset(self):
        """根据用户权限过滤任务列表"""
        # 记录用户查询参数
        logger.info(f"[TaskListView] 用户ID: {self.request.user.id}, 用户名: {self.request.user.username}")
        logger.info(f"[TaskListView] 查询参数: {dict(self.request.GET.items())}")
        
        # 获取基础查询集
        queryset = super().get_queryset()
        
        # 应用任务权限过滤
        filtered_queryset = TaskPermission.filter_tasks_by_role(queryset, self.request.user)
        
        # 处理搜索过滤
        name = self.request.GET.get('name')
        advert_name = self.request.GET.get('advert_name')
        status = self.request.GET.get('status')
        project_name = self.request.GET.get('project_name')
        optimizer = self.request.GET.get('optimizer')
        
        # 打印原始查询结果数量，用于对比
        initial_count = filtered_queryset.count()
        logger.info(f"[TaskListView] 初始任务数量: {initial_count}")
        
        if name:
            filtered_queryset = filtered_queryset.filter(name__icontains=name)
        if advert_name:
            filtered_queryset = filtered_queryset.filter(advert_name__icontains=advert_name)
        if status:
            filtered_queryset = filtered_queryset.filter(status__TaskStatusID=status)
        
        # 使用project__ProjectName进行过滤
        if project_name:
            filtered_queryset = filtered_queryset.filter(project__ProjectName__icontains=project_name)
            
        if optimizer:
            # 处理逗号分隔的optimizer IDs
            optimizer_ids = optimizer.split(',')
            filtered_queryset = filtered_queryset.filter(optimizer__id__in=optimizer_ids)
        
        # 记录过滤后的任务数量
        final_count = filtered_queryset.count()
        logger.info(f"[TaskListView] 过滤后任务数量: {final_count}")
        
        # 如果过滤后没有结果，添加提示信息
        if initial_count > 0 and filtered_queryset.count() == 0:
            logger.warning("[TaskListView] 过滤后没有结果")
            messages.warning(self.request, '没有找到符合条件的任务，请尝试调整筛选条件')
        
        # 应用排序并返回
        return filtered_queryset.distinct().order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        """添加额外上下文数据"""
        context = super().get_context_data(**kwargs)
        
        # 添加任务状态列表到上下文
        context['statuses'] = TaskStatus.objects.filter(CompanyID=self.request.user.company)
        
        # 添加当前用户可见的项目列表 - 只显示用户所在公司的项目
        context['projects'] = Project.objects.filter(CompanyID=self.request.user.company)
        
        # 添加优化师列表到上下文
        context['optimizers'] = User.objects.filter(
            Q(groups__name='部门主管') |
            Q(groups__name='小组长') |
            Q(groups__name='优化师'),
            company=self.request.user.company
        ).distinct()
        
        return context

class TaskCreateView(LoginRequiredMixin, CreateView):
    model = Task
    template_name = 'admin/task_management/task/task_form.html'
    fields = ['name', 'advert_name', 'project', 'product_info', 'status', 'backend', 
             'start_date', 'end_date', 'notes', 'pixel', 'publish_url', 'timezone', 'optimizer']
    success_url = reverse_lazy('task_management:task_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = '新建任务'
        
        # 添加项目列表到上下文 - 只显示用户所在公司的项目
        context['projects'] = Project.objects.filter(CompanyID=self.request.user.company).order_by('ProjectName')
        
        # 如果是编辑模式，获取任务实例
        task_id = self.request.GET.get('id')
        if task_id:
            try:
                # 检查权限
                permitted_tasks = TaskPermission.filter_tasks_by_role(Task.objects.filter(pk=task_id), self.request.user)
                if not permitted_tasks.exists():
                    messages.error(self.request, '您没有权限编辑此任务')
                    return context
                
                task = permitted_tasks.first()
                context['form'] = self.get_form_class()(instance=task)
                context['title'] = '编辑任务'
                context['task'] = task  # 添加任务实例到上下文
                # 获取项目相关信息
                project = task.project
                context.update({
                    'media_channel': project.MediaChannelID.MediaChannelName if project.MediaChannelID else '-',
                    'task_type': project.TaskTypeID.TaskTypeName if project.TaskTypeID else '-',
                    'daily_report_url': project.DailyReportURL or '-',
                    'kpi': project.KPI or '-',
                    'manager': project.ManagerID.username if project.ManagerID else '-',
                    'timezone': project.TimeZone or '-',
                    'status2': project.Status2.TaskStatusName if project.Status2 else '-',
                    'product_backend': project.ProductBackend or '-'
                })
            except Task.DoesNotExist:
                messages.error(self.request, '任务不存在')
                pass
        
        # 添加任务状态列表到上下文
        context['statuses'] = TaskStatus.objects.filter(CompanyID=self.request.user.company)
        
        # 添加优化师列表到上下文
        context['optimizers'] = User.objects.filter(
            Q(groups__name='部门主管') |
            Q(groups__name='小组长') |
            Q(groups__name='优化师'),
            company=self.request.user.company
        ).distinct()
        
        return context

    def form_valid(self, form):
        try:
            # 设置公司
            form.instance.company = self.request.user.company
            
            # 处理时区字段中的Unicode转义序列
            if 'timezone' in form.cleaned_data and form.cleaned_data['timezone']:
                form.instance.timezone = decode_unicode_escapes(form.cleaned_data['timezone'])
            
            # 检查是否是编辑模式
            task_id = self.request.GET.get('id')
            if task_id:
                # 编辑模式：检查除了当前任务外是否有重名任务
                existing_task = Task.objects.filter(
                    project=form.instance.project,
                    name=form.instance.name
                ).exclude(pk=task_id).first()
                if existing_task:
                    form.add_error('name', '该项目下已存在同名任务')
                    return self.form_invalid(form)
                
                # 手动保存表单数据到现有实例
                if self.object:
                    # 更新实例的每个字段
                    for field in form.cleaned_data:
                        if field != 'optimizer':  # 多对多字段需要特殊处理
                            if field == 'timezone' and form.cleaned_data[field]:
                                # 处理时区字段
                                setattr(self.object, field, decode_unicode_escapes(form.cleaned_data[field]))
                            else:
                                setattr(self.object, field, form.cleaned_data[field])
                    
                    # 保存更新后的实例
                    self.object.save()
                    
                    # 处理多对多字段
                    if 'optimizer' in form.cleaned_data:
                        self.object.optimizer.set(form.cleaned_data['optimizer'])
                    
                    # 添加成功消息
                    messages.success(self.request, '任务已成功更新')
                    
                    # 获取当前搜索参数
                    search_params = {}
                    for key in ['name', 'status', 'project_name', 'optimizer']:
                        value = self.request.GET.get(key)
                        if value:
                            search_params[key] = value
                    
                    # 构建重定向URL，包含搜索参数
                    success_url = self.get_success_url()
                    if search_params:
                        query_string = '&'.join([f"{k}={v}" for k, v in search_params.items()])
                        success_url = f"{success_url}?{query_string}"
                    
                    return HttpResponseRedirect(success_url)
            else:
                # 新建模式：检查是否有重名任务
                if Task.objects.filter(
                    project=form.instance.project,
                    name=form.instance.name
                ).exists():
                    form.add_error('name', '该项目下已存在同名任务')
                    return self.form_invalid(form)
                
                # 添加成功消息
                messages.success(self.request, '任务已成功创建')
            
            # 对于新建模式，使用默认的保存方法
            return super().form_valid(form)
        except IntegrityError:
            form.add_error('name', '保存失败，请检查任务名称是否重复')
            return self.form_invalid(form)
        except ValidationError as e:
            for field, errors in e.message_dict.items():
                form.add_error(field, errors)
            return self.form_invalid(form)

    def get_form(self, form_class=None):
        """获取表单实例，如果是编辑模式，使用现有任务实例初始化表单"""
        form_class = self.get_form_class()
        kwargs = self.get_form_kwargs()
        
        # 检查是否是编辑模式
        task_id = self.request.GET.get('id')
        if task_id and hasattr(self, 'object') and self.object and 'instance' not in kwargs:
            # 使用现有实例初始化表单，但避免重复传递instance参数
            kwargs['instance'] = self.object
        
        return form_class(**kwargs)

    def post(self, request, *args, **kwargs):
        # 检查是否是编辑模式
        task_id = request.GET.get('id')
        if task_id:
            try:
                # 获取现有任务实例
                self.object = get_object_or_404(Task, pk=task_id, company=request.user.company)
            except Task.DoesNotExist:
                messages.error(request, '任务不存在或无权限编辑')
                return HttpResponseRedirect(self.success_url)
        else:
            # 新建模式
            self.object = None
        
        # 处理表单提交
        form = self.get_form()
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

def get_project_info(request):
    """获取项目信息的AJAX视图"""
    project_id = request.GET.get('project_id')
    if project_id:
        try:
            project = Project.objects.get(pk=project_id)
            
            # 预处理可能包含特殊字符的字段
            daily_report_url = project.DailyReportURL or '-'
            kpi = project.KPI or '-'
            product_backend = project.ProductBackend or '-'
            timezone = project.TimeZone or '-'
            
            # 对时区值进行解码处理
            timezone = decode_unicode_escapes(timezone)
            
            data = {
                'success': True,
                'media_channel': project.MediaChannelID.MediaChannelName if project.MediaChannelID else '-',
                'task_type': project.TaskTypeID.TaskTypeName if project.TaskTypeID else '-',
                'daily_report_url': daily_report_url,
                'kpi': kpi,
                'manager': project.ManagerID.username if project.ManagerID else '-',
                'timezone': timezone,
                'status2': project.Status2.TaskStatusName if project.Status2 else '-',
                'product_backend': product_backend
            }
        except Project.DoesNotExist:
            data = {'success': False, 'message': '项目不存在'}
    else:
        data = {'success': False, 'message': '未提供项目ID'}
    return JsonResponse(data, json_dumps_params={'ensure_ascii': False})

class TaskDeleteView(LoginRequiredMixin, DeleteView):
    model = Task
    success_url = reverse_lazy('task_management:task_list')
    
    def get_queryset(self):
        # 确保用户只能删除自己公司的任务
        return Task.objects.filter(company=self.request.user.company)
    
    def delete(self, request, *args, **kwargs):
        try:
            self.object = self.get_object()
            success_url = self.get_success_url()
            self.object.delete()
            messages.success(request, '任务已成功删除')
            return HttpResponseRedirect(success_url)
        except Exception as e:
            messages.error(request, f'删除失败: {str(e)}')
            return HttpResponseRedirect(self.success_url)

def get_task_detail(request):
    """获取任务详情的AJAX视图"""
    task_id = request.GET.get('task_id')
    if task_id:
        try:
            # 获取任务对象
            task = Task.objects.get(pk=task_id)
            
            # 使用TaskPermission检查权限
            permitted_tasks = TaskPermission.filter_tasks_by_role(Task.objects.filter(pk=task_id), request.user)
            if not permitted_tasks.exists():
                return JsonResponse({'success': False, 'message': '您没有权限查看此任务'})
            
            # 获取优化师列表
            optimizers = [{"id": optimizer.id, "username": optimizer.username} for optimizer in task.optimizer.all()]
            
            # 预处理可能包含特殊字符的字段
            product_info = task.product_info or ''
            backend = task.backend or ''
            timezone = task.timezone or ''
            notes = task.notes or ''
            pixel = task.pixel or ''
            publish_url = task.publish_url or ''
            
            # 对时区值进行解码处理
            timezone = decode_unicode_escapes(timezone)
            
            # 项目相关字段
            daily_report_url = task.project.DailyReportURL or '-' if task.project else '-'
            kpi = task.project.KPI or '-' if task.project else '-'
            product_backend = task.project.ProductBackend or '-' if task.project else '-'
            project_timezone = task.project.TimeZone or '-' if task.project else '-'
            
            # 对项目时区值进行解码处理
            project_timezone = decode_unicode_escapes(project_timezone)
            
            data = {
                'success': True,
                'id': str(task.id),
                'name': task.name,
                'advert_name': task.advert_name,
                'project': {
                    'id': task.project.ProjectID,
                    'name': task.project.ProjectName,
                    'media_channel': task.project.MediaChannelID.MediaChannelName if task.project.MediaChannelID else '-',
                    'task_type': task.project.TaskTypeID.TaskTypeName if task.project and task.project.TaskTypeID else '-',
                    'daily_report_url': daily_report_url,
                    'kpi': kpi,
                    'manager': task.project.ManagerID.username if task.project and task.project.ManagerID else '-',
                    'timezone': project_timezone,
                    'status2': task.project.Status2.TaskStatusName if task.project.Status2 else '-',
                    'product_backend': product_backend
                },
                'product_info': product_info,
                'status': {
                    'id': task.status.TaskStatusID,
                    'name': task.status.TaskStatusName
                },
                'backend': backend,
                'timezone': timezone,
                'created_at': task.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'start_date': task.start_date.strftime('%Y-%m-%d') if task.start_date else '-',
                'end_date': task.end_date.strftime('%Y-%m-%d') if task.end_date else '-',
                'notes': notes,
                'pixel': pixel,
                'publish_url': publish_url,
                'optimizers': optimizers,
                'company_name': task.company.company_name if task.company else '-'
            }
        except Task.DoesNotExist:
            data = {'success': False, 'message': '任务不存在'}
    else:
        data = {'success': False, 'message': '未提供任务ID'}
    return JsonResponse(data, json_dumps_params={'ensure_ascii': False})

def download_task_template(request):
    """提供任务批量导入的Excel模板下载"""
    
    # 创建Excel工作簿和工作表
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "任务导入模板"
    
    # 定义字段 - 调整顺序与截图一致
    fields = [
        "项目ID*", "项目名称", "任务名称*", "广告命名*", "优化师", "时区", 
        "广告像素", "投放链接", "产品信息", "任务状态*", "产品后台", "开始日期*", "结束日期", "备注"
    ]
    
    # 设置表头样式
    header_fill = PatternFill(start_color="4361EE", end_color="4361EE", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # 写入表头
    for col_num, field in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col_num, value=field)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        # 调整列宽
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15
    
    # 添加说明行
    ws.cell(row=2, column=1, value="请填写项目ID").font = Font(italic=True)
    ws.cell(row=2, column=3, value="示例: 新年活动").font = Font(italic=True)
    ws.cell(row=2, column=4, value="示例: 2024新春促销").font = Font(italic=True)
    ws.cell(row=2, column=5, value="多个用户名用逗号分隔").font = Font(italic=True)
    ws.cell(row=2, column=6, value="示例: UTC+8").font = Font(italic=True)
    ws.cell(row=2, column=10, value="进行中/已完成/暂停").font = Font(italic=True)
    ws.cell(row=2, column=12, value="格式: YYYY-MM-DD").font = Font(italic=True)
    ws.cell(row=2, column=13, value="格式: YYYY-MM-DD").font = Font(italic=True)
    
    # 添加第二个工作表作为参考数据
    ref_ws = wb.create_sheet(title="参考数据")
    
    # 项目数据
    projects = Project.objects.filter(CompanyID=request.user.company)
    ref_ws.cell(row=1, column=1, value="项目ID").font = Font(bold=True)
    ref_ws.cell(row=1, column=2, value="项目名称").font = Font(bold=True)
    
    for i, project in enumerate(projects, 2):
        ref_ws.cell(row=i, column=1, value=str(project.ProjectID))
        ref_ws.cell(row=i, column=2, value=project.ProjectName)
    
    # 任务状态数据
    ref_ws.cell(row=1, column=4, value="可用状态").font = Font(bold=True)
    statuses = TaskStatus.objects.all()
    for i, status in enumerate(statuses, 2):
        ref_ws.cell(row=i, column=4, value=status.TaskStatusName)
    
    # 优化师数据
    ref_ws.cell(row=1, column=6, value="优化师用户名").font = Font(bold=True)
    optimizers = User.objects.filter(
        Q(groups__name='部门主管') |
        Q(groups__name='小组长') |
        Q(groups__name='优化师'),
        company=request.user.company
    ).distinct()
    for i, optimizer in enumerate(optimizers, 2):
        ref_ws.cell(row=i, column=6, value=optimizer.username)
    
    # 时区参考
    ref_ws.cell(row=1, column=8, value="常用时区").font = Font(bold=True)
    timezones = ["UTC+8", "UTC+0", "UTC-5", "UTC+1", "UTC+9"]
    for i, tz in enumerate(timezones, 2):
        ref_ws.cell(row=i, column=8, value=tz)
    
    # 调整列宽
    for col in [1, 2, 4, 6, 8]:
        ref_ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    # 设置响应头
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=task_import_template.xlsx'
    
    # 保存工作簿到响应
    wb.save(response)
    return response

def export_tasks_excel(request):
    """将任务数据导出为Excel文件"""
    # 获取筛选条件
    name = request.GET.get('name', '')
    advert_name = request.GET.get('advert_name', '')
    status = request.GET.get('status', '')
    project = request.GET.get('project', '')
    optimizer = request.GET.get('optimizer', '')
    
    # 获取基础查询集
    tasks = Task.objects.all()
    
    # 应用任务权限过滤
    tasks = TaskPermission.filter_tasks_by_role(tasks, request.user)
    
    # 应用筛选条件
    if name:
        tasks = tasks.filter(name__icontains=name)
    if advert_name:
        tasks = tasks.filter(advert_name__icontains=advert_name)
    if status:
        tasks = tasks.filter(status__TaskStatusID=status)
    if project:
        tasks = tasks.filter(project__ProjectID=project)
    if optimizer:
        optimizer_ids = optimizer.split(',')
        tasks = tasks.filter(optimizer__id__in=optimizer_ids)
    
    # 创建Excel工作簿和工作表
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "任务列表"
    
    # 设置表头样式
    header_fill = PatternFill(start_color="4361EE", end_color="4361EE", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # 定义表头
    headers = [
        "序号", "任务名称", "广告命名", "所属项目", "媒体渠道", "任务类型", 
        "任务状态", "KPI", "时区", "日报链接", "广告像素", "投放链接", 
        "优化师", "运营负责人", "开始日期", "结束日期"
    ]
    
    # 写入表头
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        # 调整列宽
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15
    
    # 写入数据
    for row_num, task in enumerate(tasks, 2):
        # 序号
        ws.cell(row=row_num, column=1, value=row_num-1)
        # 任务名称
        ws.cell(row=row_num, column=2, value=task.name)
        # 广告命名
        ws.cell(row=row_num, column=3, value=task.advert_name)
        # 所属项目
        ws.cell(row=row_num, column=4, value=task.project.ProjectName if task.project else '')
        # 媒体渠道
        ws.cell(row=row_num, column=5, value=task.project.MediaChannelID.MediaChannelName if task.project and task.project.MediaChannelID else '')
        # 任务类型
        ws.cell(row=row_num, column=6, value=task.project.TaskTypeID.TaskTypeName if task.project and task.project.TaskTypeID else '')
        # 任务状态
        ws.cell(row=row_num, column=7, value=task.status.TaskStatusName if task.status else '')
        # KPI
        ws.cell(row=row_num, column=8, value=task.project.KPI if task.project else '')
        # 时区
        ws.cell(row=row_num, column=9, value=task.timezone or '')
        # 日报链接
        ws.cell(row=row_num, column=10, value=task.project.DailyReportURL if task.project else '')
        # 广告像素
        ws.cell(row=row_num, column=11, value=task.pixel or '')
        # 投放链接
        ws.cell(row=row_num, column=12, value=task.publish_url or '')
        # 优化师
        optimizers = ', '.join([optimizer.username for optimizer in task.optimizer.all()])
        ws.cell(row=row_num, column=13, value=optimizers)
        # 运营负责人
        ws.cell(row=row_num, column=14, value=task.project.ManagerID.username if task.project and task.project.ManagerID else '')
        # 开始日期
        ws.cell(row=row_num, column=15, value=task.start_date)
        # 结束日期
        ws.cell(row=row_num, column=16, value=task.end_date if task.end_date else '')
    
    # 设置响应头
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=tasks_export_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx'
    
    # 保存工作簿到响应
    wb.save(response)
    return response

def import_tasks(request):
    """批量导入任务"""
    if request.method == 'POST':
        file = request.FILES.get('task_file')
        
        if not file:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'status': 'error', 'message': '请选择要上传的Excel文件'})
            messages.error(request, '请选择要上传的Excel文件')
            return redirect('task_management:task_list')
        
        if not file.name.endswith(('.xlsx', '.xls')):
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'status': 'error', 'message': '请上传Excel格式的文件（.xlsx或.xls）'})
            messages.error(request, '请上传Excel格式的文件（.xlsx或.xls）')
            return redirect('task_management:task_list')
        
        try:
            # 读取Excel文件 - 强制所有列为字符串类型以避免pandas自动类型转换
            df = pd.read_excel(file, sheet_name=0, dtype=str)
            print(f"开始导入: {df.shape[0]}行数据")
            
            # 清洗数据 - 去除列名中的空格和隐藏字符
            df.columns = [col.strip() for col in df.columns]
            
            # 验证必要列
            required_columns = ['项目ID*', '任务名称*', '广告命名*', '任务状态*', '开始日期*']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                print(f"缺少必要列: {missing_columns}")
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'status': 'error', 
                        'message': f'Excel文件缺少必要的列：{", ".join(missing_columns)}'
                    })
                messages.error(request, f'Excel文件缺少必要的列：{", ".join(missing_columns)}')
                return redirect('task_management:task_list')
            
            # 获取任务总数用于计算进度
            total_tasks = len(df)
            
            # 开始批量导入
            success_count = 0
            error_count = 0
            update_count = 0
            skip_count = 0
            error_messages = []
            
            # 先收集所有任务名称及其对应的实例，用于检查重复
            all_tasks = {task.name: task for task in Task.objects.filter(company=request.user.company)}
            
            for index, row in df.iterrows():
                # 计算当前进度百分比
                current_progress = int((index + 1) / total_tasks * 100)
                
                # 每处理10条记录或最后一条时发送进度更新
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest' and (index % 10 == 0 or index == total_tasks - 1):
                    progress_data = {
                        'status': 'processing',
                        'progress': current_progress,
                        'current': index + 1,
                        'total': total_tasks,
                        'success': success_count,
                        'error': error_count,
                        'update': update_count,
                        'skip': skip_count
                    }
                    print(f"进度: {current_progress}% - 新增:{success_count} 更新:{update_count} 跳过:{skip_count} 错误:{error_count}")
                
                try:
                    # 检查任务名称是否已存在
                    task_name = str(row['任务名称*']).strip()
                    existing_task = all_tasks.get(task_name)
                    
                    # 获取用户选择的处理方式，默认为更新
                    duplicate_action = request.POST.get('duplicate_action', 'update')
                    
                    if existing_task and duplicate_action == 'skip':
                        # 用户选择跳过已存在任务
                        skip_count += 1
                        continue
                    
                    # 获取关联对象
                    try:
                        project_id = row['项目ID*']
                        
                        # 尝试转换项目ID为整数（处理各种格式）
                        try:
                            # 移除非数字字符
                            project_id_clean = ''.join(c for c in str(project_id) if c.isdigit())
                            if project_id_clean:
                                project_id = int(project_id_clean)
                        except (ValueError, TypeError):
                            pass
                            
                        project = Project.objects.get(ProjectID=project_id, CompanyID=request.user.company)
                    except Project.DoesNotExist:
                        error_count += 1
                        error_msg = f"第{index+2}行：项目ID {project_id} 不存在或无权访问"
                        error_messages.append(error_msg)
                        continue
                    except (ValueError, TypeError) as e:
                        error_count += 1
                        error_msg = f"第{index+2}行：项目ID格式错误 - {str(e)}"
                        error_messages.append(error_msg)
                        continue
                    
                    # 获取任务状态
                    status_name = str(row.get('任务状态*', '进行中')).strip()
                    try:
                        # 尝试精确匹配
                        status = TaskStatus.objects.get(TaskStatusName=status_name)
                    except TaskStatus.DoesNotExist:
                        # 尝试模糊匹配
                        try:
                            status = TaskStatus.objects.filter(
                                TaskStatusName__icontains=status_name
                            ).first()
                            if not status:
                                status = TaskStatus.objects.get(TaskStatusName='进行中')
                        except TaskStatus.DoesNotExist:
                            error_count += 1
                            error_msg = f"第{index+2}行：未找到默认任务状态'进行中'"
                            error_messages.append(error_msg)
                            continue
                    
                    # 处理日期格式
                    try:
                        start_date_value = row['开始日期*']
                        
                        # 解析开始日期
                        start_date = parse_date_value(start_date_value)
                        if not start_date:
                            raise ValueError(f"无法解析日期格式: {start_date_value}")
                            
                        # 处理结束日期（可选）
                        end_date = None
                        if '结束日期' in row and not pd.isna(row['结束日期']):
                            end_date_value = row['结束日期']
                            end_date = parse_date_value(end_date_value)
                    except ValueError as e:
                        error_count += 1
                        error_msg = f"第{index+2}行：日期格式错误 - {str(e)}"
                        error_messages.append(error_msg)
                        continue
                    
                    # 判断是更新还是创建新任务
                    if existing_task:
                        # 更新已存在的任务
                        task = existing_task
                        task.advert_name = str(row['广告命名*'])
                        task.project = project
                        task.status = status
                        task.start_date = start_date
                        task.end_date = end_date
                        
                        # 更新可选字段
                        if '产品信息' in row and not pd.isna(row['产品信息']):
                            task.product_info = str(row['产品信息'])
                        if '产品后台' in row and not pd.isna(row['产品后台']):
                            task.backend = str(row['产品后台'])
                        if '备注' in row and not pd.isna(row['备注']):
                            task.notes = str(row['备注'])
                        if '广告像素' in row and not pd.isna(row['广告像素']):
                            task.pixel = str(row['广告像素'])
                        if '投放链接' in row and not pd.isna(row['投放链接']):
                            task.publish_url = str(row['投放链接'])
                        if '时区' in row and not pd.isna(row['时区']):
                            # 解码时区值中可能的Unicode转义序列
                            timezone_value = str(row['时区'])
                            # 将Unicode转义序列解码为实际字符
                            task.timezone = decode_unicode_escapes(timezone_value)
                            
                        task.save()
                        update_count += 1
                    else:
                        # 创建新任务
                        task = Task(
                            name=task_name,
                            advert_name=str(row['广告命名*']),
                            project=project,
                            status=status,
                            start_date=start_date,
                            end_date=end_date,
                            company=request.user.company,
                            # 可选字段
                            product_info=str(row['产品信息']) if '产品信息' in row and not pd.isna(row['产品信息']) else '',
                            backend=str(row['产品后台']) if '产品后台' in row and not pd.isna(row['产品后台']) else '',
                            notes=str(row['备注']) if '备注' in row and not pd.isna(row['备注']) else '',
                            pixel=str(row['广告像素']) if '广告像素' in row and not pd.isna(row['广告像素']) else '',
                            publish_url=str(row['投放链接']) if '投放链接' in row and not pd.isna(row['投放链接']) else '',
                            timezone=decode_unicode_escapes(str(row['时区'])) if '时区' in row and not pd.isna(row['时区']) else '',
                        )
                        task.save()
                        
                        # 将任务添加到已存在的任务集合中
                        all_tasks[task.name] = task
                        success_count += 1
                    
                    # 处理优化师关联（可选）- 清除现有关联再重新添加
                    if '优化师' in row and not pd.isna(row['优化师']):
                        # 如果是更新操作，先清除现有的优化师关联
                        if existing_task:
                            task.optimizer.clear()
                            
                        optimizer_names = str(row['优化师']).split(',')
                        
                        for name in optimizer_names:
                            name = name.strip()
                            if not name:
                                continue
                                
                            try:
                                user = User.objects.get(username=name, company=request.user.company)
                                task.optimizer.add(user)
                            except User.DoesNotExist:
                                continue
                    
                    # 处理媒体渠道关联（可选）
                    if '媒体渠道' in row and not pd.isna(row['媒体渠道']):
                        media_channel_name = str(row['媒体渠道']).strip()
                        
                        try:
                            from task_management.models import MediaChannel
                            channel = MediaChannel.objects.filter(
                                MediaChannelName__icontains=media_channel_name
                            ).first()
                            
                            if channel:
                                task.media_channel = channel
                                task.save()
                        except Exception:
                            pass
                    
                except Exception as e:
                    error_count += 1
                    error_msg = f"第{index+2}行处理异常: {str(e)}"
                    error_messages.append(error_msg)
            
            # 处理完成，返回最终结果
            result = {
                'status': 'complete',
                'progress': 100,
                'success': success_count,
                'error': error_count,
                'update': update_count,
                'skip': skip_count,
                'total': total_tasks
            }
            
            # 显示导入结果
            print(f"导入完成: 共{total_tasks}条 - 新增:{success_count} 更新:{update_count} 跳过:{skip_count} 错误:{error_count}")
            
            if success_count > 0:
                success_msg = f'成功导入 {success_count} 个新任务'
                result['success_message'] = success_msg
                messages.success(request, success_msg)
            
            if update_count > 0:
                update_msg = f'成功更新 {update_count} 个已存在的任务'
                result['update_message'] = update_msg
                messages.info(request, update_msg)
            
            if skip_count > 0:
                skip_msg = f'跳过 {skip_count} 个已存在的任务'
                result['skip_message'] = skip_msg
                messages.warning(request, skip_msg)
            
            if error_count > 0:
                error_summary = f'导入失败 {error_count} 个任务'
                result['error_message'] = error_summary
                
                if error_messages:
                    # 限制错误消息数量，避免页面过长
                    if len(error_messages) > 5:
                        error_detail = '<br>'.join(error_messages[:5]) + f'<br>... 等{len(error_messages)}个错误'
                    else:
                        error_detail = '<br>'.join(error_messages)
                    error_summary += f'<br>{error_detail}'
                    result['error_detail'] = error_messages[:5] if len(error_messages) > 5 else error_messages
                messages.error(request, error_summary)
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse(result)
            return redirect('task_management:task_list')
            
        except Exception as e:
            error_msg = f'文件处理失败: {str(e)}'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'status': 'error', 'message': error_msg})
            messages.error(request, error_msg)
            return redirect('task_management:task_list')
    
    # GET请求显示上传表单
    return render(request, 'admin/task_management/task/task_import.html')

def parse_date_value(date_value):
    """解析各种格式的日期值，返回date对象"""
    if pd.isna(date_value):
        return None
        
    # 转为字符串处理
    date_str = str(date_value).strip()
    
    # Excel日期数值格式 (数字)
    try:
        # 尝试将字符串转换为浮点数
        if date_str.replace('.', '', 1).isdigit():
            excel_serial = float(date_str)
            if 1 <= excel_serial <= 100000:  # 合理的Excel日期范围
                date_obj = datetime(1899, 12, 30) + timedelta(days=excel_serial)
                return date_obj.date()
    except (ValueError, OverflowError):
        pass
    
    # 清理日期字符串，移除非法字符
    date_clean = ''.join(c for c in date_str if c.isdigit() or c in ['-', '/', '.'])
    
    # 检查特殊情况 - "20205/3/31" 这种格式
    if len(date_clean) > 10 and ('/' in date_clean or '-' in date_clean or '.' in date_clean):
        # 确定分隔符
        sep = '/' if '/' in date_clean else ('-' if '-' in date_clean else '.')
        parts = date_clean.split(sep)
        
        if len(parts) >= 2 and len(parts[0]) > 4:
            # 可能是年份多写了数字，例如"20205"应该是"2020"
            year = parts[0][:4]  # 取前4位作为年份
            month = parts[1]
            day = parts[2] if len(parts) > 2 else "1"
            
            corrected_date = f"{year}{sep}{month}{sep}{day}"
            
            try:
                if sep == '/':
                    date_obj = datetime.strptime(corrected_date, '%Y/%m/%d').date()
                elif sep == '-':
                    date_obj = datetime.strptime(corrected_date, '%Y-%m-%d').date()
                else:
                    date_obj = datetime.strptime(corrected_date, '%Y.%m.%d').date()
                return date_obj
            except ValueError:
                pass
    
    # 尝试常见日期格式
    date_formats = [
        '%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d',  # 标准格式
        '%m/%d/%Y', '%d/%m/%Y',              # 美式/英式格式
        '%y-%m-%d', '%y/%m/%d',              # 短年份格式
        '%Y%m%d', '%y%m%d'                   # 无分隔符格式
    ]
    
    for fmt in date_formats:
        try:
            date_obj = datetime.strptime(date_clean, fmt).date()
            return date_obj
        except ValueError:
            continue
    
    # 尝试从当前日期解析年月日
    if date_clean.isdigit():
        try:
            # YYYYMMDD 格式
            if len(date_clean) == 8:
                year = int(date_clean[:4])
                month = int(date_clean[4:6])
                day = int(date_clean[6:8])
                if 1 <= month <= 12 and 1 <= day <= 31:
                    return datetime(year, month, day).date()
            
            # YYMMDD 格式
            elif len(date_clean) == 6:
                year = 2000 + int(date_clean[:2])  # 假设是21世纪
                month = int(date_clean[2:4])
                day = int(date_clean[4:6])
                if 1 <= month <= 12 and 1 <= day <= 31:
                    return datetime(year, month, day).date()
        except (ValueError, OverflowError):
            pass
            
    # 尝试直接通过pandas处理
    try:
        date_obj = pd.to_datetime(date_value).date()
        return date_obj
    except:
        pass
        
    return None

def update_task_status(request):
    """
    API视图：更新任务状态
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': '只支持POST请求'}, status=405)
    
    task_id = request.POST.get('task_id')
    status_id = request.POST.get('status_id')
    
    if not task_id or not status_id:
        return JsonResponse({'success': False, 'error': '缺少必要参数'}, status=400)
    
    try:
        # 检查权限
        permitted_tasks = TaskPermission.filter_tasks_by_role(Task.objects.filter(pk=task_id), request.user)
        if not permitted_tasks.exists():
            return JsonResponse({'success': False, 'error': '您没有权限编辑此任务'}, status=403)
        
        task = permitted_tasks.first()
        status = TaskStatus.objects.get(TaskStatusID=status_id)
        
        # 更新任务状态
        task.status = status
        task.save(update_fields=['status'])
        
        return JsonResponse({
            'success': True,
            'status_name': status.TaskStatusName,
            'message': f'成功更新任务状态为"{status.TaskStatusName}"'
        })
    except Task.DoesNotExist:
        return JsonResponse({'success': False, 'error': '任务不存在'}, status=404)
    except TaskStatus.DoesNotExist:
        return JsonResponse({'success': False, 'error': '状态不存在'}, status=404)
    except Exception as e:
        logger.error(f"更新任务状态时出错: {str(e)}")
        return JsonResponse({'success': False, 'error': f'更新失败: {str(e)}'}, status=500)

def update_task_end_date(request):
    """
    API视图：更新任务结束日期
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': '只支持POST请求'}, status=405)
    
    task_id = request.POST.get('task_id')
    end_date = request.POST.get('end_date')
    
    if not task_id or not end_date:
        return JsonResponse({'success': False, 'error': '缺少必要参数'}, status=400)
    
    try:
        # 检查权限
        permitted_tasks = TaskPermission.filter_tasks_by_role(Task.objects.filter(pk=task_id), request.user)
        if not permitted_tasks.exists():
            return JsonResponse({'success': False, 'error': '您没有权限编辑此任务'}, status=403)
        
        task = permitted_tasks.first()
        
        # 验证日期格式
        try:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'success': False, 'error': '日期格式无效，应为YYYY-MM-DD'}, status=400)
        
        # 验证结束日期必须晚于开始日期
        if task.start_date and end_date_obj < task.start_date:
            return JsonResponse({'success': False, 'error': '结束日期必须晚于开始日期'}, status=400)
        
        # 更新任务结束日期
        task.end_date = end_date_obj
        task.save(update_fields=['end_date'])
        
        return JsonResponse({
            'success': True,
            'end_date': end_date,
            'message': f'成功更新任务结束日期为"{end_date}"'
        })
    except Task.DoesNotExist:
        return JsonResponse({'success': False, 'error': '任务不存在'}, status=404)
    except Exception as e:
        logger.error(f"更新任务结束日期时出错: {str(e)}")
        return JsonResponse({'success': False, 'error': f'更新失败: {str(e)}'}, status=500)